import io
import re
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook

from scenario_models import ScenarioRecord


def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[\s\-_\/:()]+", "", s)
    return s


COLUMN_ALIASES = {
    "scenario_id": ["scnid", "scenarioid", "id"],
    "original_id": ["origid", "originalid"],
    "module": ["module", "domain", "area"],
    "focus_area": ["focusarea", "focus", "submodule"],
    "scenario_type": ["scenariotype", "type", "executiontype"],
    "scenario_title": ["scenariotitle", "title", "testname", "name"],
    "test_objective": ["testobjective", "objective", "expectedbehaviour", "expectedbehavior"],
    "priority": ["priority", "severity"],
    "test_type": ["testtype", "positivenegative", "typeclass"],
    "status": ["status", "executionstatus"],
    "remarks": ["remarks", "notes", "comment"],
    "user_query": ["userquery", "query", "prompt", "initialmessage", "userprompt"],
    "expected_response": ["expectedresponse", "expectedanswer", "expectedoutput"],
    "source_kb": ["sourcekb", "kb", "source"],
    "action": ["action", "expectedaction"],
    "tool_calling": ["toolcalling", "toolcallingqueries", "tool", "toolquery"],
    "preconditions": ["preconditions", "precondition"],
    "dependencies": ["dependencies", "dependson", "dependency"],
    "language": ["language"],
    "domain": ["domain"],
    "requires_attachment": ["requiresattachment", "attachmentrequired"],
    "requires_ticket": ["requiresticket", "ticketrequired"],
    "requires_card_interaction": ["requirescardinteraction", "cardinteractionrequired"],
    "requires_admin_access": ["requiresadminaccess", "adminrequired"],
}


def _cell(row, idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    val = row[idx]
    return str(val).strip() if val is not None else ""


def _to_bool(val: str) -> bool:
    return (val or "").strip().lower() in ("y", "yes", "true", "1", "t")


def _split_multi(val: str) -> List[str]:
    if not val:
        return []
    parts = re.split(r"[,\n;|]+", val)
    return [p.strip() for p in parts if p.strip()]


def _find_header_map(headers: List[str]) -> Dict[str, int]:
    norm_headers = [_norm(h) for h in headers]
    mapping: Dict[str, int] = {}

    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in norm_headers:
                mapping[canonical] = norm_headers.index(alias)
                break

    return mapping


def _score_header_row(headers: List[str]) -> int:
    """
    Score a row for how likely it is to be a header row.
    """
    header_map = _find_header_map(headers)
    score = len(header_map)

    # Strong indicators
    strong = 0
    if "scenario_id" in header_map:
        strong += 2
    if "module" in header_map:
        strong += 1
    if "scenario_title" in header_map:
        strong += 2
    if "test_objective" in header_map:
        strong += 2
    if "priority" in header_map:
        strong += 1
    if "status" in header_map:
        strong += 1

    return score + strong


def _find_best_header_row(ws, scan_rows: int = 15) -> Tuple[Optional[int], Dict[str, int], List[str]]:
    """
    Scan first N rows and find the most likely header row.
    """
    best_row_idx = None
    best_score = 0
    best_map = {}
    best_headers = []

    max_scan = min(scan_rows, ws.max_row)
    for r in range(1, max_scan + 1):
        vals = [str(c.value or "").strip() for c in ws[r]]
        if not any(vals):
            continue

        score = _score_header_row(vals)
        if score > best_score:
            best_score = score
            best_row_idx = r
            best_map = _find_header_map(vals)
            best_headers = vals

    # Require at least some meaningful structure
    if best_score < 2:
        return None, {}, []

    return best_row_idx, best_map, best_headers


def _merge_record_fields(rec: ScenarioRecord, row_vals: list, header_map: dict, headers: list, sheet_name: str):
    field_map = [
        ("original_id", "original_id"),
        ("module", "module"),
        ("focus_area", "focus_area"),
        ("scenario_type", "scenario_type"),
        ("scenario_title", "scenario_title"),
        ("test_objective", "test_objective"),
        ("priority", "priority"),
        ("test_type", "test_type"),
        ("status", "status"),
        ("remarks", "remarks"),
        ("user_query", "user_query"),
        ("expected_response", "expected_response"),
        ("source_kb", "source_kb"),
        ("action", "action"),
        ("language", "language"),
        ("domain", "domain"),
    ]

    for field_name, hdr_key in field_map:
        current = getattr(rec, field_name)
        incoming = _cell(row_vals, header_map.get(hdr_key, -1))
        if (not current) and incoming:
            setattr(rec, field_name, incoming)

    rec.tool_calling = rec.tool_calling or _to_bool(_cell(row_vals, header_map.get("tool_calling", -1)))
    rec.requires_attachment = rec.requires_attachment or _to_bool(_cell(row_vals, header_map.get("requires_attachment", -1)))
    rec.requires_ticket = rec.requires_ticket or _to_bool(_cell(row_vals, header_map.get("requires_ticket", -1)))
    rec.requires_card_interaction = rec.requires_card_interaction or _to_bool(_cell(row_vals, header_map.get("requires_card_interaction", -1)))
    rec.requires_admin_access = rec.requires_admin_access or _to_bool(_cell(row_vals, header_map.get("requires_admin_access", -1)))

    rec.preconditions = list(dict.fromkeys(rec.preconditions + _split_multi(_cell(row_vals, header_map.get("preconditions", -1)))))
    rec.dependencies = list(dict.fromkeys(rec.dependencies + _split_multi(_cell(row_vals, header_map.get("dependencies", -1)))))

    raw_map = {headers[i]: (row_vals[i] if i < len(row_vals) else "") for i in range(len(headers))}
    rec.raw.setdefault("source_sheets", []).append(sheet_name)
    rec.raw.setdefault("rows", []).append(raw_map)


def _is_probable_data_row(row_vals: list, header_map: dict) -> bool:
    scenario_id = _cell(row_vals, header_map.get("scenario_id", -1))
    title = _cell(row_vals, header_map.get("scenario_title", -1))
    objective = _cell(row_vals, header_map.get("test_objective", -1))
    module = _cell(row_vals, header_map.get("module", -1))

    # Any useful content
    if any([scenario_id, title, objective, module]):
        return True

    return False


def _looks_like_scn_id(value: str) -> bool:
    v = (value or "").strip().upper()
    return bool(re.match(r"SCN[-\s]?\d+", v))


def load_structured_workbook(filename: str, raw: bytes) -> Dict:
    if not raw:
        return {"suite_name": filename, "records": [], "errors": ["Empty file"], "sheet_summaries": []}

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        return {"suite_name": filename, "records": [], "errors": [f"Failed to parse workbook: {e}"], "sheet_summaries": []}

    records: Dict[str, ScenarioRecord] = {}
    errors: List[str] = []
    sheet_summaries: List[Dict] = []

    for ws in wb.worksheets:
        header_row_idx, header_map, headers = _find_best_header_row(ws, scan_rows=20)

        if not header_row_idx or not header_map:
            sheet_summaries.append({"sheet": ws.title, "rows": 0, "used": False, "header_row": None})
            continue

        used_count = 0
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            row_vals = list(row)

            if not _is_probable_data_row(row_vals, header_map):
                continue

            scenario_id = _cell(row_vals, header_map.get("scenario_id", -1))
            title = _cell(row_vals, header_map.get("scenario_title", -1))
            objective = _cell(row_vals, header_map.get("test_objective", -1))
            module = _cell(row_vals, header_map.get("module", -1))

            # Avoid accidental title/section rows
            if not any([scenario_id, title, objective, module]):
                continue

            # Better keying
            key = None
            if scenario_id and _looks_like_scn_id(scenario_id):
                key = scenario_id.strip()
            elif title:
                key = f"TITLE::{title.strip()}"
            elif objective:
                key = f"OBJ::{objective.strip()[:80]}"
            else:
                continue

            if key not in records:
                rec = ScenarioRecord(
                    scenario_id=scenario_id or title or f"ROW-{len(records)+1:03d}",
                    original_id=_cell(row_vals, header_map.get("original_id", -1)),
                    module=module,
                    focus_area=_cell(row_vals, header_map.get("focus_area", -1)),
                    scenario_type=_cell(row_vals, header_map.get("scenario_type", -1)),
                    scenario_title=title,
                    test_objective=objective,
                    priority=_cell(row_vals, header_map.get("priority", -1)) or "medium",
                    test_type=_cell(row_vals, header_map.get("test_type", -1)),
                    status=_cell(row_vals, header_map.get("status", -1)) or "Not Tested",
                    remarks=_cell(row_vals, header_map.get("remarks", -1)),
                    user_query=_cell(row_vals, header_map.get("user_query", -1)),
                    expected_response=_cell(row_vals, header_map.get("expected_response", -1)),
                    source_kb=_cell(row_vals, header_map.get("source_kb", -1)),
                    action=_cell(row_vals, header_map.get("action", -1)),
                    tool_calling=_to_bool(_cell(row_vals, header_map.get("tool_calling", -1))),
                    preconditions=_split_multi(_cell(row_vals, header_map.get("preconditions", -1))),
                    dependencies=_split_multi(_cell(row_vals, header_map.get("dependencies", -1))),
                    language=_cell(row_vals, header_map.get("language", -1)) or "English",
                    domain=_cell(row_vals, header_map.get("domain", -1)),
                    requires_attachment=_to_bool(_cell(row_vals, header_map.get("requires_attachment", -1))),
                    requires_ticket=_to_bool(_cell(row_vals, header_map.get("requires_ticket", -1))),
                    requires_card_interaction=_to_bool(_cell(row_vals, header_map.get("requires_card_interaction", -1))),
                    requires_admin_access=_to_bool(_cell(row_vals, header_map.get("requires_admin_access", -1))),
                    raw={"source_sheets": [], "rows": []},
                )
                _merge_record_fields(rec, row_vals, header_map, headers, ws.title)
                records[key] = rec
            else:
                _merge_record_fields(records[key], row_vals, header_map, headers, ws.title)

            used_count += 1

        sheet_summaries.append({
            "sheet": ws.title,
            "rows": used_count,
            "used": used_count > 0,
            "header_row": header_row_idx,
        })

    out = list(records.values())

    # Extra fallback: if we still got nothing, try sheet-name-specific heuristics later
    if not out:
        errors.append("No structured scenarios could be normalized from workbook")

    return {
        "suite_name": filename,
        "records": out,
        "errors": errors,
        "sheet_summaries": sheet_summaries,
    }
