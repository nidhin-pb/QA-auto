"""
QA Report Generator (2 sheets)

Sheet 1: QA Summary
- Clean, tester-friendly presentation
- Key results, tickets, and links

Sheet 2: Failure Details  
- Only failed/error tests
- Conversation transcript + links for debugging

No circular imports (does not import TestResult).
"""
import os
import re
from datetime import datetime
from typing import List, Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import app_config
from utils import ensure_dir, timestamp, extract_all_ticket_numbers


class ReportGenerator:
    def __init__(self):
        ensure_dir(app_config.report_dir)

    def generate_excel_report(self, results: List[Any]) -> str:
        return self.generate_report(results)

    def generate_report(self, results: List[Any]) -> str:
        if not results:
            return ""

        wb = Workbook()
        wb.remove(wb.active)

        self._create_qa_summary_sheet(wb, results)
        self._create_compact_results_sheet(wb, results)
        self._create_test_cases_reference_sheet(wb, results)
        self._create_failure_evidence_sheet(wb, results)

        filename = f"CVA_QA_Report_{timestamp()}.xlsx"
        filepath = os.path.join(app_config.report_dir, filename)
        wb.save(filepath)
        return filepath

    # ---------------- Helpers ----------------

    def _get_excel_meta(self, r: Any) -> Dict:
        scenario = getattr(r, "scenario", {}) or {}
        excel = scenario.get("excel", {}) or {}
        return excel

    def _get_first_bot_reply(self, r: Any) -> str:
        for m in (getattr(r, "conversation_log", []) or []):
            role = (m.get("role") or "").lower()
            if role in ("assistant", "cva"):
                return (m.get("content") or "").strip()
        return ""

    def _get_all_bot_text(self, r: Any) -> str:
        parts = []
        for m in (getattr(r, "conversation_log", []) or []):
            role = (m.get("role") or "").lower()
            if role in ("assistant", "cva"):
                parts.append(m.get("content") or "")
        return "\n\n".join(parts)

    def _get_all_links(self, r: Any) -> List[str]:
        links = []
        for m in (getattr(r, "conversation_log", []) or []):
            for u in (m.get("links") or []):
                if u and u not in links:
                    links.append(u)
        return links

    def _normalize(self, s: str) -> str:
        s = (s or "").lower()
        s = re.sub(r"[\s_]+", "", s)
        return s.strip()

    def _kb_links_only(self, links: List[str]) -> List[str]:
        out = []
        for u in links or []:
            lu = (u or "").lower()
            if "knowledgebasestaging.blob.core.windows.net" in lu or lu.endswith(".pdf") or ".pdf" in lu or ".docx" in lu:
                out.append(u)
        return out

    def _action_detected_summary(self, r: Any) -> str:
        """
        A short human-friendly summary of what CVA actually did.
        """
        text = self._get_all_bot_text(r)
        t = (text or "").lower()
        links = self._get_all_links(r)
        tickets = extract_all_ticket_numbers(text or "")
        has_ticket = len(tickets) > 0
        has_catalog = ("complete this request" in t) or ("service catalog" in t)
        has_steps = ("step" in t) or ("1." in t) or ("try" in t) or ("check" in t)
        has_servicenow_link = any("service-now.com" in (u or "").lower() or "servicenow" in (u or "").lower() for u in links)

        bits = []
        if has_steps:
            bits.append("Troubleshooting steps")
        if has_catalog:
            bits.append("Catalog/options shown")
        if has_ticket:
            bits.append(f"Ticket found: {', '.join(tickets[:3])}")
        if has_servicenow_link and not has_ticket:
            bits.append("ServiceNow link shown")
        
        # Add ticket lifecycle state information
        state = getattr(r, "state", {})
        if getattr(r, "state", {}).get("ticket_created"):
            bits.append("Ticket lifecycle: Created")
        if getattr(r, "state", {}).get("ticket_updated"):
            bits.append("Ticket lifecycle: Updated")
        if getattr(r, "state", {}).get("ticket_resolved"):
            bits.append("Ticket lifecycle: Resolved")

        if not bits:
            bits.append("General response only")

        family = ((getattr(r, "scenario", {}) or {}).get("excel", {}) or {}).get("family", "")
        if family:
            bits.append(f"Family: {family}")

        return " | ".join(bits)

    def _top_failure_reason(self, r: Any) -> str:
        bugs = getattr(r, "bugs_found", None) or []
        if bugs:
            return str(bugs[0])[:800]

        notes = getattr(r, "notes", "") or ""
        if notes.strip():
            return notes[:800]

        alt = getattr(r, "alternate_reason", "") or ""
        if alt.strip():
            return alt[:800]

        goal = getattr(r, "goal_achieved_reason", "") or ""
        if goal.strip():
            return goal[:800]

        vfail = getattr(r, "validations_failed", None) or []
        if vfail:
            return f"Failed validations: {', '.join(vfail[:5])}"

        err = getattr(r, "error_message", "") or ""
        if err.strip():
            return err[:800]

        return ""

    # ---------------- Sheet 1 ----------------

    def _create_qa_summary_sheet(self, wb: Workbook, results: List[Any]):
        ws = wb.create_sheet("QA Summary", 0)

        headers = [
            "Test ID",
            "Scenario Title",
            "Module",
            "Family",
            "Execution Mode",
            "User Prompt",
            "Actual Bot Response",
            "QA Status",
            "QA Score",
            "QA Grade",
            "Key Result / Reason",
            "Ticket ID(s)",
            "KB Link(s)",
        ]

        header_fill = PatternFill(start_color="2F855A", end_color="2F855A", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="top")

        for row_idx, r in enumerate(results, start=2):
            all_bot = self._get_all_bot_text(r)
            first_bot = self._get_first_bot_reply(r)
            excel = self._get_excel_meta(r)
            tickets = extract_all_ticket_numbers(all_bot)
            links = self._get_all_links(r)

            family = getattr(r, "structured_family", "") or excel.get("family", "")
            key_reason = self._top_failure_reason(r) or getattr(r, "goal_achieved_reason", "") or getattr(r, "notes", "")

            ws.cell(row=row_idx, column=1, value=getattr(r, "test_id", ""))
            ws.cell(row=row_idx, column=2, value=getattr(r, "test_name", ""))
            ws.cell(row=row_idx, column=3, value=getattr(r, "category", ""))
            ws.cell(row=row_idx, column=4, value=family)
            ws.cell(row=row_idx, column=5, value=getattr(r, "execution_mode", ""))
            ws.cell(row=row_idx, column=6, value=(excel.get("user_query") or "")[:1500])
            ws.cell(row=row_idx, column=7, value=(first_bot or "")[:3000])
            ws.cell(row=row_idx, column=8, value=getattr(r, "display_status", getattr(r, "final_status", getattr(r, "status", ""))))
            ws.cell(row=row_idx, column=9, value=getattr(r, "qa_score", ""))
            ws.cell(row=row_idx, column=10, value=getattr(r, "qa_grade", ""))
            ws.cell(row=row_idx, column=11, value=(key_reason or "")[:1500])
            ws.cell(row=row_idx, column=12, value=", ".join(tickets[:10]))
            ws.cell(row=row_idx, column=13, value="\n".join(links[:10]))

            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

        widths = [12, 40, 24, 18, 18, 40, 50, 16, 10, 12, 45, 22, 40]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------- Sheet 2 ----------------

    def _create_compact_results_sheet(self, wb: Workbook, results: List[Any]):
        ws = wb.create_sheet("QA Results (Compact)", 1)

        headers = [
            "Test ID",
            "Scenario Title",
            "Module",
            "Family",
            "Execution Mode",
            "Priority",
            "User Prompt",
            "First Bot Reply",
            "Action Detected",
            "QA Status",
            "QA Score",
            "QA Grade",
            "Ticket ID(s)",
            "KB / Links",
            "Goal Achieved",
            "Key Reason / Notes",
            "Alternate Outcome",
            "Conversation Turns",
            "Duration (s)",
        ]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="top")

        for row_idx, r in enumerate(results, start=2):
            excel = self._get_excel_meta(r)
            first_bot = self._get_first_bot_reply(r)
            all_bot = self._get_all_bot_text(r)
            links = self._get_all_links(r)
            tickets = extract_all_ticket_numbers(all_bot)
            conv_log = getattr(r, "conversation_log", []) or []
            num_turns = len([m for m in conv_log if (m.get("role") or "").lower() in ("assistant", "cva")])

            family = getattr(r, "structured_family", "") or excel.get("family", "")
            key_reason = self._top_failure_reason(r) or getattr(r, "goal_achieved_reason", "") or getattr(r, "notes", "")

            ws.cell(row=row_idx, column=1, value=getattr(r, "test_id", ""))
            ws.cell(row=row_idx, column=2, value=getattr(r, "test_name", ""))
            ws.cell(row=row_idx, column=3, value=getattr(r, "category", ""))
            ws.cell(row=row_idx, column=4, value=family)
            ws.cell(row=row_idx, column=5, value=getattr(r, "execution_mode", ""))
            ws.cell(row=row_idx, column=6, value=getattr(r, "priority", ""))
            ws.cell(row=row_idx, column=7, value=(excel.get("user_query") or "")[:1500])
            ws.cell(row=row_idx, column=8, value=(first_bot or "")[:3000])
            ws.cell(row=row_idx, column=9, value=self._action_detected_summary(r))
            ws.cell(row=row_idx, column=10, value=getattr(r, "display_status", getattr(r, "final_status", getattr(r, "status", ""))))
            ws.cell(row=row_idx, column=11, value=getattr(r, "qa_score", ""))
            ws.cell(row=row_idx, column=12, value=getattr(r, "qa_grade", ""))
            ws.cell(row=row_idx, column=13, value=", ".join(tickets[:10]))
            ws.cell(row=row_idx, column=14, value="\n".join(links[:5]))
            ws.cell(row=row_idx, column=15, value=getattr(r, "goal_achieved_reason", "") or "")
            ws.cell(row=row_idx, column=16, value=(key_reason or "")[:1500])
            ws.cell(row=row_idx, column=17, value="Yes" if getattr(r, "alternate_outcome", False) else "No")
            ws.cell(row=row_idx, column=18, value=num_turns)
            ws.cell(row=row_idx, column=19, value=round(getattr(r, "duration", 0), 1))

            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

        widths = [12, 45, 25, 18, 18, 10, 45, 55, 30, 16, 10, 12, 22, 40, 30, 45, 14, 12, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------- Sheet 3 ----------------

    def _create_test_cases_reference_sheet(self, wb: Workbook, results: List[Any]):
        """Sheet 4: Shows detailed test cases from the Excel for reference."""
        ws = wb.create_sheet("Test Cases Reference")

        headers = [
            "Scenario ID",
            "Scenario Title",
            "Test Objective",
            "Expected Response",
            "Source KB",
            "Action",
            "Tool Calling",
            "Automation Status",
            "QA Result",
            "QA Score",
        ]

        header_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="top")

        # Build a lookup from results
        result_map = {}
        for r in results:
            result_map[getattr(r, "test_id", "")] = r

        row_idx = 2
        for r in results:
            excel = self._get_excel_meta(r)
            test_id = getattr(r, "test_id", "")

            # Get raw data which may contain multiple test case rows
            scenario = getattr(r, "scenario", {}) or {}
            raw = ((scenario.get("excel", {}) or {}).get("raw", {}) or {})
            raw_rows = raw.get("rows", []) if isinstance(raw, dict) else []

            if raw_rows and len(raw_rows) > 1:
                # Multiple test case rows merged — show each
                for raw_row in raw_rows:
                    ws.cell(row=row_idx, column=1, value=test_id)
                    ws.cell(row=row_idx, column=2, value=getattr(r, "test_name", ""))

                    # Extract fields from raw row
                    obj = ""
                    exp = ""
                    src = ""
                    action = ""
                    tool = ""
                    for key, val in raw_row.items():
                        kl = (key or "").lower().strip()
                        val_str = str(val).strip() if val else ""
                        if "objective" in kl or "expected behaviour" in kl:
                            obj = val_str or obj
                        if "expected" in kl and "response" in kl:
                            exp = val_str or exp
                        if "source" in kl or "kb" in kl:
                            src = val_str or src
                        if "action" in kl:
                            action = val_str or action
                        if "tool" in kl:
                            tool = val_str or tool

                    ws.cell(row=row_idx, column=3, value=(obj or excel.get("action", ""))[:2000])
                    ws.cell(row=row_idx, column=4, value=(exp or excel.get("expected_response", ""))[:3000])
                    ws.cell(row=row_idx, column=5, value=(src or excel.get("source_kb", ""))[:500])
                    ws.cell(row=row_idx, column=6, value=(action or excel.get("action", ""))[:1000])
                    ws.cell(row=row_idx, column=7, value="Y" if excel.get("tool_calling_queries") else "N")
                    ws.cell(row=row_idx, column=8, value=getattr(r, "execution_mode", ""))
                    ws.cell(row=row_idx, column=9, value=getattr(r, "display_status", ""))
                    ws.cell(row=row_idx, column=10, value=getattr(r, "qa_score", ""))

                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
                    row_idx += 1
            else:
                # Single row — just show the scenario
                ws.cell(row=row_idx, column=1, value=test_id)
                ws.cell(row=row_idx, column=2, value=getattr(r, "test_name", ""))
                ws.cell(row=row_idx, column=3, value=(excel.get("action", "") or excel.get("test_objective", ""))[:2000])
                ws.cell(row=row_idx, column=4, value=(excel.get("expected_response", ""))[:3000])
                ws.cell(row=row_idx, column=5, value=(excel.get("source_kb", ""))[:500])
                ws.cell(row=row_idx, column=6, value=(excel.get("action", ""))[:1000])
                ws.cell(row=row_idx, column=7, value="Y" if excel.get("tool_calling_queries") else "N")
                ws.cell(row=row_idx, column=8, value=getattr(r, "execution_mode", ""))
                ws.cell(row=row_idx, column=9, value=getattr(r, "display_status", ""))
                ws.cell(row=row_idx, column=10, value=getattr(r, "qa_score", ""))

                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
                row_idx += 1

        widths = [12, 45, 50, 55, 30, 35, 12, 18, 14, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ---------------- Sheet 4 ----------------

    def _create_failure_evidence_sheet(self, wb: Workbook, results: List[Any]):
        ws = wb.create_sheet("Failure Details")

        headers = [
            "Test ID", "Status",
            "Client", "Module",
            "User Query",
            "Action (Expected)",
            "Source KB (Expected)",
            "Top Bug / Failure Reason",
            "Conversation Transcript",
            "Captured Links",
        ]

        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="top")

        row_idx = 2
        for r in results:
            status = getattr(r, "status", "")
            if status not in ("failed", "error"):
                continue

            excel = self._get_excel_meta(r)
            links = self._get_all_links(r)

            # Transcript
            transcript_lines = []
            for m in (getattr(r, "conversation_log", []) or []):
                role = (m.get("role") or "").lower()
                who = "USER" if role == "user" else "CVA"
                ts = m.get("timestamp", "")
                content = (m.get("content") or "").strip()
                transcript_lines.append(f"[{who}] {ts}\n{content}\n")
            transcript = "\n".join(transcript_lines)[:20000]

            ws.cell(row=row_idx, column=1, value=getattr(r, "test_id", ""))
            ws.cell(row=row_idx, column=2, value=status)
            ws.cell(row=row_idx, column=3, value=excel.get("client", ""))
            ws.cell(row=row_idx, column=4, value=excel.get("module", ""))
            ws.cell(row=row_idx, column=5, value=(excel.get("user_query") or "")[:3000])
            ws.cell(row=row_idx, column=6, value=(excel.get("action") or "")[:1200])
            ws.cell(row=row_idx, column=7, value=(excel.get("source_kb") or "")[:800])
            ws.cell(row=row_idx, column=8, value=self._top_failure_reason(r))
            ws.cell(row=row_idx, column=9, value=transcript)
            ws.cell(row=row_idx, column=10, value="\n".join(links)[:8000])

            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

            row_idx += 1

        widths = [10, 10, 14, 16, 42, 30, 26, 55, 80, 55]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
