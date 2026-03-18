import io
import re
import csv
from typing import List, Dict, Tuple, Optional
from openpyxl import load_workbook
from intent import Intent


def _norm_header(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[\s\-_]+", "", s)
    return s


def map_intent(action: str, module: str):
    a = (action or "").lower()
    m = (module or "").lower()

    if "greeting" in m:
        return Intent.GREETING

    if "out-of-scope" in m:
        return Intent.OUT_OF_SCOPE

    if "catalog" in m:
        return Intent.CATALOG

    if "network" in m:
        return Intent.TROUBLESHOOT

    if "create" in a:
        return Intent.CREATE_TICKET

    if "update" in a:
        return Intent.UPDATE_TICKET

    if "status" in a:
        return Intent.STATUS_CHECK

    if "close" in a:
        return Intent.CLOSE_TICKET

    if "reopen" in a:
        return Intent.REOPEN_TICKET

    return Intent.UNKNOWN


def _is_greeting(q: str) -> bool:
    qn = (q or "").strip().lower()
    # Remove trailing punctuation for matching
    qn_clean = re.sub(r"[!?,.\s]+$", "", qn).strip()
    
    exact = {
        "hi", "hello", "hey", "hey there", "hi / hello",
        "good morning", "good afternoon", "good evening", "morning",
        "howdy", "greetings", "yo",
    }
    if qn_clean in exact:
        return True
    
    # Partial greeting patterns
    greeting_starts = [
        "hi,", "hi ", "hello,", "hello ", "hey,", "hey ",
        "good morning", "good afternoon", "good evening",
    ]
    if any(qn.startswith(p) for p in greeting_starts):
        return True

    return False


def _is_closing_or_thanks(q: str) -> bool:
    qn = (q or "").strip().lower()
    return any(x in qn for x in [
        "thank you", "thanks", "appreciate", "that resolved", "resolved my issue",
        "problem fixed", "all sorted", "issue resolved", "that's all",
        "that resolved my issue", "thanks!", "thank you for your help",
    ])


def _is_off_topic(q: str) -> bool:
    qn = (q or "").strip().lower()
    return any(x in qn for x in [
        "prime minister", "match tonight", "who will win", "cricket", "football",
        "weather like", "weather today", "movie", "celebrity", "politics",
        "reimburse", "travel expenses", "reimbursement",
        "order lunch", "order food", "lunch for me",
        "what's the weather", "who is the president",
        "tell me a joke", "sing a song", "play music",
    ])


def _is_single_turn_from_content(q: str) -> bool:
    """Detect if query is inherently single-turn based on content."""
    return _is_greeting(q) or _is_closing_or_thanks(q) or _is_off_topic(q)


def _to_bool(val) -> bool:
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ("y", "yes", "true", "1", "t")


def _parse_query_type(val) -> str:
    """Parse Query Type column: 'Single-Turn' or 'Multi-Turn'."""
    s = (str(val) if val else "").strip().lower()
    s = re.sub(r"[\s\-_]+", "", s)
    if "single" in s:
        return "single"
    if "multi" in s:
        return "multi"
    return ""  # unknown / not specified


def _guess_turns(action: str, tool_calling: bool, query_type: str, user_query: str) -> Tuple[int, int]:
    """
    Determine min/max turns.
    Priority:
      1. Excel Query Type column (Single-Turn / Multi-Turn)
      2. Content-based detection (greeting, off-topic, thanks)
      3. Action-based heuristic
    """
    # If Excel explicitly says Single-Turn
    if query_type == "single":
        return 1, 1

    # If content is inherently single-turn
    if _is_single_turn_from_content(user_query):
        return 1, 1

    # If Excel explicitly says Multi-Turn
    if query_type == "multi":
        a = (action or "").lower()
        if tool_calling or any(x in a for x in ["ticket", "incident", "service request", "catalog", "draft"]):
            return 2, 6
        if any(x in a for x in ["troubleshoot", "trouble", "provide troubleshooting", "steps"]):
            return 3, 7
        return 2, 5

    # Fallback: action-based heuristic
    a = (action or "").lower()
    if any(x in a for x in ["hi", "hello", "greet"]):
        return 1, 1
    if tool_calling or any(x in a for x in ["ticket", "incident", "service request", "catalog", "draft"]):
        return 2, 6
    if any(x in a for x in ["troubleshoot", "trouble", "provide troubleshooting", "steps"]):
        return 3, 7
    return 2, 5


def _build_validations(action: str, expected_response: str, source_kb: str, tool_calling: bool) -> List[str]:
    validations = []
    a = (action or "").lower()
    src = (source_kb or "").strip()
    exp = (expected_response or "").strip()

    if ("troubleshoot" in a) or ("steps" in a) or (len(src) > 0):
        validations.append("provides_troubleshooting_steps")
        validations.append("includes_kb_hyperlink")
        if src:
            validations.append("includes_specific_kb")

    if tool_calling:
        validations.append("has_tool_action")

    if exp:
        validations.append("matches_expected_semantic")

    if src:
        validations.append("does_not_contradict_kb_availability")

    validations.append("handles_gracefully")
    return validations


def load_suite_from_bytes(filename: str, raw: bytes) -> Dict:
    suite_name = filename
    errors: List[str] = []
    cases: List[Dict] = []

    if not raw:
        return {"suite_name": suite_name, "cases": [], "errors": ["Empty file"]}

    lower = (filename or "").lower()
    try:
        if lower.endswith(".csv"):
            rows = _read_csv(raw)
        else:
            rows = _read_xlsx(raw)
    except Exception as e:
        return {"suite_name": suite_name, "cases": [], "errors": [f"Failed to parse file: {e}"]}

    if not rows:
        return {"suite_name": suite_name, "cases": [], "errors": ["No rows found"]}

    for idx, r in enumerate(rows, start=1):
        user_query = (r.get("user_query") or "").strip()
        if not user_query:
            continue

        client = (r.get("client") or "").strip()
        module = (r.get("module") or "").strip() or "Excel Suite"
        category = (r.get("category") or "").strip()
        query_type_raw = (r.get("query_type") or "").strip()
        difficulty = (r.get("difficulty") or "").strip()
        expected_response = (r.get("expected_response") or "").strip()
        source_kb = (r.get("source_kb") or "").strip()
        action = (r.get("action") or "").strip()
        tool_calling = _to_bool(r.get("tool_calling_queries"))

        query_type = _parse_query_type(query_type_raw)

        min_turns, max_turns = _guess_turns(action, tool_calling, query_type, user_query)
        validations = _build_validations(action, expected_response, source_kb, tool_calling)

        # Determine if single-turn
        single_turn = (min_turns == 1 and max_turns == 1)

        response_timeout = 90
        if "agent" in action.lower() or "transfer" in action.lower():
            response_timeout = 140

        scenario = {
            "id": f"XL-{idx:03d}",
            "name": f"[{module}] {user_query[:60]}",
            "category": category or module,
            "priority": "high" if tool_calling else "medium",
            "goal": action or "Validate response",
            "description": f"Client={client} Module={module} QueryType={query_type_raw} Difficulty={difficulty} ToolCalling={tool_calling}",
            "initial_message": user_query,
            "min_turns": min_turns,
            "max_turns": max_turns,
            "validations": validations,
            "stop_after_first_response": single_turn,
            "response_timeout": response_timeout,
            "intent": map_intent(action, module),

            "excel": {
                "client": client,
                "module": module,
                "category": category,
                "query_type": query_type_raw,
                "difficulty": difficulty,
                "expected_response": expected_response,
                "source_kb": source_kb,
                "action": action,
                "tool_calling_queries": tool_calling,
                "user_query": user_query,
            }
        }
        cases.append(scenario)

    if not cases:
        errors.append("No valid test cases (missing User Query?)")

    return {"suite_name": suite_name, "cases": cases, "errors": errors}


def _read_xlsx(raw: bytes) -> List[Dict]:
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(_norm_header(str(cell.value or "")))

    col_map = _map_columns(headers)

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for key, col_idx in col_map.items():
            if col_idx < len(row):
                val = row[col_idx]
                d[key] = str(val).strip() if val is not None else ""
        out.append(d)
    return out


def _read_csv(raw: bytes) -> List[Dict]:
    text = raw.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(io.StringIO(text))
    field_map = {_norm_header(h): h for h in (reader.fieldnames or [])}
    headers_norm = list(field_map.keys())
    col_map = _map_columns_csv(headers_norm, field_map)

    out = []
    for r in reader:
        d = {}
        for key, original_header in col_map.items():
            d[key] = (r.get(original_header, "") or "").strip()
        out.append(d)
    return out


def _map_columns(headers_norm: List[str]) -> Dict[str, int]:
    """For XLSX: returns {key: column_index}."""
    candidates = {
        "client": ["client"],
        "module": ["module"],
        "category": ["category"],
        "query_type": ["querytype", "querytypes", "type", "turntype"],
        "difficulty": ["difficulty", "difficultylevel"],
        "user_query": ["userquery", "userquestion", "query", "userprompt"],
        "expected_response": ["expectedresponse", "expectedanswer", "expected"],
        "source_kb": ["sourcekb", "source", "kb", "kbsource"],
        "action": ["action", "expectedaction"],
        "tool_calling_queries": ["toolcallingqueries", "toolcalling", "tool", "toolqueries"],
    }

    mapping: Dict[str, int] = {}
    for key, opts in candidates.items():
        for opt in opts:
            if opt in headers_norm:
                mapping[key] = headers_norm.index(opt)
                break

    if "user_query" not in mapping:
        raise ValueError("Missing required column: User Query")

    return mapping


def _map_columns_csv(headers_norm: List[str], field_map: Dict[str, str]) -> Dict[str, str]:
    """For CSV: returns {key: original_header_name}."""
    candidates = {
        "client": ["client"],
        "module": ["module"],
        "category": ["category"],
        "query_type": ["querytype", "querytypes", "type", "turntype"],
        "difficulty": ["difficulty", "difficultylevel"],
        "user_query": ["userquery", "userquestion", "query", "userprompt"],
        "expected_response": ["expectedresponse", "expectedanswer", "expected"],
        "source_kb": ["sourcekb", "source", "kb", "kbsource"],
        "action": ["action", "expectedaction"],
        "tool_calling_queries": ["toolcallingqueries", "toolcalling", "tool", "toolqueries"],
    }

    mapping: Dict[str, str] = {}
    for key, opts in candidates.items():
        for opt in opts:
            if opt in field_map:
                mapping[key] = field_map[opt]
                break

    if "user_query" not in mapping:
        raise ValueError("Missing required column: User Query")

    return mapping
